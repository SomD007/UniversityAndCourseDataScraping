import time
import re
import json
import random
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# configuration
UNI_LIST_URL   = "https://www.mastersportal.com/search/universities/master"
COURSES_SEARCH = "https://www.mastersportal.com/search/master/"  # + ?organisationId=X part will be here
MAX_UNIVERSITIES = 25
COURSES_PER_UNI  = 5
HEADLESS   = False  # so we can see the browser window
SLOW_MO    = 80
PAGE_WAIT  = 3.5 # seconds to wait after page load
MIN_DELAY  = 4.0 # min delay between requests
MAX_DELAY  = 7.0 # max delay between requests
OUTPUT_FILE = "university_course_data.xlsx"


# HELPERS

def clean(text) -> str:
    # collapse whitespace and return 'N/A' for empty values
    if not text:
        return "N/A"
    return re.sub(r"\s+", " ", str(text).strip()) or "N/A"


def human_delay(lo=MIN_DELAY, hi=MAX_DELAY):
     # random sleep to mimic human browsing and avoid bot detection
    t = random.uniform(lo, hi)
    print(f"  Waiting {t:.1f}s ...")
    time.sleep(t)


def print_divider(char="─", width=65):
     # adds a line separator between sections in the terminal output
    print(char * width)


def extract_org_id(url: str):
     # pull numeric org ID from a MastersPortal university URL
    m = re.search(r"/universities/(\d+)/", url)
    return m.group(1) if m else None


# scrape university listing page
def scrape_university_list(page) -> list:
    print(f"\n Loading: {UNI_LIST_URL}")
    page.goto(UNI_LIST_URL, wait_until="domcontentloaded", timeout=45000)
    time.sleep(PAGE_WAIT)

    # scroll to trigger lazy loading cards
    for _ in range(3):
        page.keyboard.press("End")
        time.sleep(1.5)

    # find card container
    cards = None
    for sel in [".OrganisationCard", ".UniversityCard", "[class*='OrganisationCard']"]:
        cnt = page.locator(sel).count()
        if cnt > 0:
            print(f"  Found {cnt} university cards  [{sel}]")
            cards = page.locator(sel)
            break

    if not cards:
        print("No cards found. Page text:")
        print(page.inner_text("body")[:2000])
        return []

    universities = []
    total = min(cards.count(), MAX_UNIVERSITIES)

    for i in range(total):
        card = cards.nth(i)

        data = card.evaluate("""el => {
            const nameEl = el.querySelector('h2.OrganisationName, h2, h3');
            const name   = nameEl ? nameEl.innerText.trim() : 'N/A';

            let location = 'N/A';
            el.querySelectorAll('div.Fact').forEach(fact => {
                const lbl = fact.querySelector('div.Label');
                const val = fact.querySelector('div.Value');
                if (lbl && val && lbl.innerText.trim() === 'Location')
                    location = val.innerText.trim();
            });

            // the <a> tag wraps the whole card - this gives us the mastersportal link
            const cardA = el.closest('a') || el.querySelector('a[href]');
            const href  = cardA ? cardA.getAttribute('href') : 'N/A';

            // official university website sits inside the "Visit University Page" button
            // in the card footer - confirmed from DevTools inspection
            let website = 'N/A';
            const visitBtn = el.querySelector(
                'span.VisitOrganisationButton, [class*="VisitOrganisation"], [class*="VisitUniversity"]'
            );
            if (visitBtn) {
                const visitA = visitBtn.closest('a') || visitBtn.parentElement?.closest('a');
                if (visitA) {
                    const raw = visitA.getAttribute('href') || '';
                    try { website = new URL(raw).origin + new URL(raw).pathname; }
                    catch(e) { website = raw.split('?')[0]; }
                }
            }

            // fallback: any footer link that is not a mastersportal.com link
            if (website === 'N/A') {
                const footer = el.querySelector('footer, [class*="CardFooter"]');
                if (footer) {
                    footer.querySelectorAll('a[href]').forEach(a => {
                        const h = a.getAttribute('href') || '';
                        if (h.startsWith('http') && !h.includes('mastersportal') && website === 'N/A') {
                            try { website = new URL(h).origin + new URL(h).pathname; }
                            catch(e) { website = h.split('?')[0]; }
                        }
                    });
                }
            }

            return { name, location, href, website };
        }""")

        name    = clean(data.get("name",     "N/A"))
        href    = data.get("href",    "N/A") or "N/A"
        loc_raw = clean(data.get("location", "N/A"))
        website = clean(data.get("website",  "N/A"))

        # location comes back as a combined string e.g. "Madrid, Spain"
        # split it into city and country
        parts   = [p.strip() for p in loc_raw.split(",")]
        city    = parts[0]  if len(parts) >= 1 else "N/A"
        country = parts[-1] if len(parts) >= 2 else "N/A"
        if city == country:
            country = "N/A"

        if "multiple" in loc_raw.lower():
            print(f"  [{i+1}] Skipping '{name}' — multiple locations")
            continue

        if href.startswith("/"):
            detail_url = "https://www.mastersportal.com" + href
        elif href.startswith("http"):
            detail_url = href
        else:
            detail_url = "N/A"

        org_id = extract_org_id(detail_url)

        universities.append({
            "university_name": name,
            "city":       city,
            "country":    country,
            "website":    website,
            "detail_url": detail_url,
            "org_id":     org_id,
        })
        print(f"  [{i+1}] {name} | {city}, {country} | website={website}")

    return universities


# scraping courses using /search/master/?organisationId=X

COURSE_JS = """el => {
    function factVal(el, keywords) {
        let result = 'N/A';
        el.querySelectorAll('[class*="Fact"], div.Fact').forEach(f => {
            const lbl = f.querySelector('[class*="Label"], div.Label');
            const val = f.querySelector('[class*="Value"], div.Value');
            if (lbl && val) {
                const t = lbl.innerText.trim().toLowerCase();
                if (keywords.some(k => t.includes(k)))
                    result = val.innerText.trim();
            }
        });
        return result;
    }
    function fallback(el, selectors) {
        for (const s of selectors) {
            const e = el.querySelector(s);
            if (e) return e.innerText.trim();
        }
        return 'N/A';
    }

    const nameEl = el.querySelector('h2,h3,h4,[class*="Name" i],[class*="Title" i]');
    const name   = nameEl ? nameEl.innerText.trim() : 'N/A';

    let level = factVal(el, ['degree','level']);
    if (level === 'N/A') level = fallback(el,
        ['[class*="Degree" i]','[class*="Level" i]','[class*="Type" i]']);

    let discipline = factVal(el, ['discipline','field','subject']);
    if (discipline === 'N/A') discipline = fallback(el,
        ['[class*="Discipline" i]','[class*="Field" i]','[class*="Subject" i]']);

    let duration = factVal(el, ['duration','length']);
    if (duration === 'N/A') duration = fallback(el,
        ['[class*="Duration" i]','[class*="Length" i]']);

    let fees = factVal(el, ['tuition','fee','cost']);
    if (fees === 'N/A') fees = fallback(el,
        ['[class*="Tuition" i]','[class*="Fee" i]','[class*="Cost" i]']);

    let eligibility = factVal(el, ['eligib','require','entry','admission']);
    if (eligibility === 'N/A') eligibility = fallback(el,
        ['[class*="Eligib" i]','[class*="Require" i]','[class*="Entry" i]']);

    return { name, level, discipline, duration, fees, eligibility };
}"""


def scrape_courses_via_search(page, org_id, uni_name):
    if not org_id:
        print("   No org_id — skipping")
        return [], "N/A"

    url = f"{COURSES_SEARCH}?organisationId={org_id}"
    print(f" {url}")

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=35000)
        time.sleep(PAGE_WAIT)

        body_text = page.inner_text("body")
        if "you have been blocked" in body_text.lower():
            print("Cloudflare blocked — skipping this university")
            return [], "N/A"

        for _ in range(2):
            page.keyboard.press("End")
            time.sleep(1.2)

        # search for course card containers
        items, matched_sel = None, None
        for sel in [
            ".ProgrammeItem", ".StudyOption", ".ResultItem",
            "[class*='ProgrammeItem']", "[class*='StudyOption']", "[class*='ResultItem']",
            ".SearchResult", "article",
        ]:
            cnt = page.locator(sel).count()
            if cnt > 0:
                print(f"     Course selector: '{sel}' ({cnt} items)")
                items, matched_sel = page.locator(sel), sel
                break

        if not items:
            print("      No course items found. Page preview:")
            print("      ", body_text[:300])
            return [], "N/A"

        # debug - shows first item HTML so selectors can be checked if needed
        try:
            html = page.locator(matched_sel).first.inner_html(timeout=3000)
            print(f"     First course item HTML:\n       {html[:500]}")
        except Exception:
            pass

        # extracting the course data
        courses = []
        for i in range(min(items.count(), COURSES_PER_UNI)):
            data = items.nth(i).evaluate(COURSE_JS)
            courses.append({
                "course_name":  clean(data.get("name",        "N/A")),
                "level":        clean(data.get("level",       "Master's")),
                "discipline":   clean(data.get("discipline",  "N/A")),
                "duration":     clean(data.get("duration",    "N/A")),
                "fees":         clean(data.get("fees",        "N/A")),
                "eligibility":  clean(data.get("eligibility", "N/A")),
            })

        return courses

    except PWTimeout:
        print("    Timeout")
        return []
    except Exception as e:
        print(f"      Error: {e}")
        return []


# exporting into excel sheet

def export_to_excel(all_universities: list, all_courses: list):
    uni_df = pd.DataFrame(all_universities, columns=[
        "university_id", "university_name", "country", "city", "website"
    ])
    crs_df = pd.DataFrame(all_courses, columns=[
        "course_id", "university_id", "course_name", "level",
        "discipline", "duration", "fees", "eligibility"
    ])

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        uni_df.to_excel(writer, sheet_name="Universities", index=False)
        crs_df.to_excel(writer, sheet_name="Courses",      index=False)

    wb = load_workbook(OUTPUT_FILE)

    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    CEL_FONT = Font(name="Arial", size=10)
    ALT_FILL = PatternFill("solid", start_color="D6E4F0")
    BS       = Side(style="thin", color="BFBFBF")
    BORDER   = Border(left=BS, right=BS, top=BS, bottom=BS)
    CENTER   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def style_sheet(ws, col_widths):
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for r, row in enumerate(ws.iter_rows(), 1):
            for cell in row:
                cell.border = BORDER
                if r == 1:
                    cell.fill      = HDR_FILL
                    cell.font      = HDR_FONT
                    cell.alignment = CENTER
                else:
                    cell.font      = CEL_FONT
                    cell.alignment = LEFT
                    if r % 2 == 0:
                        cell.fill = ALT_FILL
            ws.row_dimensions[r].height = 30 if r == 1 else 22
        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

    style_sheet(wb["Universities"], col_widths=[15, 46, 20, 18, 46])
    style_sheet(wb["Courses"],      col_widths=[12, 15, 46, 14, 28, 14, 24, 46])

    wb.save(OUTPUT_FILE)
    print(f"\n Excel file saved -> {OUTPUT_FILE}")
    print(f"   Sheet 'Universities' : {len(all_universities)} rows")
    print(f"   Sheet 'Courses'      : {len(all_courses)} rows")


# MAIN

def main():
    all_universities = []
    all_courses      = []

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=HEADLESS,
            slow_mo=SLOW_MO,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
            locale="en-US",
            timezone_id="Europe/London",
        )

        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
            Object.defineProperty(navigator, 'plugins',   { get: () => [1, 2, 3] });
            Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
        """)
        page = context.new_page()

        # getting university list
        universities = scrape_university_list(page)
        if not universities:
            print("\n No universities scraped. Exiting.")
            browser.close()
            return

        # courses for each university
        course_id = 1
        uni_id    = 1

        for uni in universities:
            print_divider()
            print(f"  [{uni_id}/{len(universities)}] {uni['university_name']}")
            print(f"   Location : {uni['city']}, {uni['country']}")
            print(f"   Org ID   : {uni['org_id']}")

            human_delay()  # delay to avoid triggering Cloudflare

            courses = scrape_courses_via_search(
                page, uni["org_id"], uni["university_name"]
            )

            all_universities.append({
                "university_id":   uni_id,
                "university_name": uni["university_name"],
                "country":         uni["country"],
                "city":            uni["city"],
                "website":         uni["website"],  # taken from the card on the listing page
            })

            print(f"   Website  : {uni['website']}")
            print(f"   Courses  : {len(courses)} found")

            for c in courses:
                c["course_id"]     = course_id
                c["university_id"] = uni_id
                all_courses.append(c)
                print(f"\n      Course #{course_id}")
                print(f"        Name        : {c['course_name']}")
                print(f"        Level       : {c['level']}")
                print(f"        Discipline  : {c['discipline']}")
                print(f"        Duration    : {c['duration']}")
                print(f"        Fees        : {c['fees']}")
                print(f"        Eligibility : {c['eligibility']}")
                course_id += 1

            uni_id += 1

        browser.close()

    print_divider("═")
    print(" SCRAPING COMPLETE")
    print_divider("═")
    print(f"  Universities : {len(all_universities)}")
    print(f"  Courses      : {len(all_courses)}")
    print_divider("═")

    print("\n UNIVERSITIES TABLE")
    print_divider()
    print(f"{'ID':<4} {'Name':<45} {'Country':<20} {'City':<18} Website")
    print_divider()
    for u in all_universities:
        print(f"{u['university_id']:<4} {u['university_name'][:44]:<45} "
              f"{u['country'][:19]:<20} {u['city'][:17]:<18} {u['website']}")

    print("\n COURSES TABLE")
    print_divider()
    print(f"{'CID':<5}{'UID':<5}{'Course Name':<38}{'Level':<14}{'Discipline':<22}"
          f"{'Duration':<12}{'Fees':<20}Eligibility")
    print_divider()
    for c in all_courses:
        print(f"{c['course_id']:<5}{c['university_id']:<5}"
              f"{c['course_name'][:37]:<38}{c['level'][:13]:<14}"
              f"{c['discipline'][:21]:<22}{c['duration'][:11]:<12}"
              f"{c['fees'][:19]:<20}{c['eligibility'][:55]}")

    print("\n RAW JSON")
    print_divider()
    print(json.dumps({"universities": all_universities, "courses": all_courses},
                     indent=2, ensure_ascii=False))

    if all_universities or all_courses:
        export_to_excel(all_universities, all_courses)
    else:
        print("\n No data collected — Excel file not created.")


if __name__ == "__main__":
    main()